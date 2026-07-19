from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import re
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

app = Flask(__name__)
app.config['SECRET_KEY'] = 'tu_clave_secreta_aqui'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///productos.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Funciones de validación
def validar_contraseña(password):
    """Valida que la contraseña tenga al menos 8 caracteres y al menos 1 número"""
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    
    if not re.search(r'\d', password):
        return False, "La contraseña debe contener al menos 1 número"
    
    return True, "Contraseña válida"

def verificar_contraseñas(password, confirm_password):
    """Verifica que las contraseñas coincidan"""
    if password != confirm_password:
        return False, "Las contraseñas no coinciden"
    return True, "Contraseñas coinciden"

# Funciones de autenticación y autorización
def login_required(f):
    """Decorador para requerir login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Debes iniciar sesión para acceder a esta página', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(required_roles):
    """Decorador para requerir roles específicos"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Debes iniciar sesión para acceder a esta página', 'error')
                return redirect(url_for('login'))
            
            user = Usuario.query.get(session['user_id'])
            if not user or user.categoria not in required_roles:
                flash('No tienes permisos para acceder a esta página', 'error')
                return redirect(url_for('index'))
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_current_user():
    """Obtiene el usuario actual de la sesión"""
    if 'user_id' in session:
        return Usuario.query.get(session['user_id'])
    return None

@app.context_processor
def inject_current_user():
    """Inyecta current_user en todos los templates"""
    return dict(current_user=get_current_user())

# Modelo de Producto
class Producto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    tipo_material = db.Column(db.String(100), nullable=False)
    precio = db.Column(db.Float, nullable=False, default=1000.0)  # Precio por unidad
    descripcion = db.Column(db.Text, nullable=True)
    proveedor = db.Column(db.String(100), nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Producto {self.nombre}>'

# Modelo de Usuario
class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    categoria = db.Column(db.String(20), nullable=False)  # usuario_comun, supervisor, gerente, cliente
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relación con Cliente (uno a uno)
    cliente = db.relationship('Cliente', backref='usuario', uselist=False, cascade='all, delete-orphan')

    def set_password(self, password):
        """Establece la contraseña hasheada usando contraseña+ID"""
        # Para usuarios nuevos, guardar la contraseña temporalmente
        # Se actualizará después de asignar el ID
        if not self.id:
            self._temp_password = password
            # Asignar un hash temporal para evitar el error de NOT NULL
            self.password_hash = generate_password_hash(password)
        else:
            # Para usuarios existentes, usar contraseña+ID
            combined = password + str(self.id)
            self.password_hash = generate_password_hash(combined)

    def check_password(self, password):
        """Verifica la contraseña"""
        if self.id:
            combined = password + str(self.id)
            return check_password_hash(self.password_hash, combined)
        return False

    def finalize_password(self):
        """Finaliza el hash de la contraseña después de asignar el ID"""
        if hasattr(self, '_temp_password') and self.id:
            combined = self._temp_password + str(self.id)
            self.password_hash = generate_password_hash(combined)
            delattr(self, '_temp_password')
            return True
        return False

    def __repr__(self):
        return f'<Usuario {self.nombre}>'

# Modelo de Cliente
class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    identificador = db.Column(db.String(50), unique=True, nullable=False)
    tipo_persona = db.Column(db.String(20), nullable=False)  # fisica o juridica
    nombre = db.Column(db.String(200), nullable=False)
    direccion = db.Column(db.String(300), nullable=True)
    telefono = db.Column(db.String(20), nullable=True)
    mail = db.Column(db.String(120), nullable=True)
    datos_adicionales = db.Column(db.Text, nullable=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relación con Usuario (uno a uno)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), unique=True, nullable=True)

    def __repr__(self):
        return f'<Cliente {self.nombre}>'

# Modelo de Orden de Compra
class OrdenCompra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    identificador = db.Column(db.Integer, unique=True, nullable=False)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    estado = db.Column(db.String(30), nullable=False, default='pendiente_confirmacion')  # pendiente_confirmacion, confirmada, en_transito, completada, cancelada
    direccion_entrega = db.Column(db.String(300), nullable=True)
    
    # Relaciones
    cliente = db.relationship('Cliente', backref='ordenes')
    
    def calcular_total(self):
        """Calcula el total de la orden"""
        total = 0
        items = ItemOrden.query.filter_by(orden_id=self.id).all()
        for item in items:
            # Respetar el precio unitario definido en cada producto
            precio_unitario = item.producto.precio if item.producto and item.producto.precio is not None else 0
            total += precio_unitario * item.cantidad
        return total
    
    def __repr__(self):
        return f'<OrdenCompra {self.identificador}>'

# Modelo ItemOrden para manejar la relación muchos a muchos con cantidad
class ItemOrden(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    orden_id = db.Column(db.Integer, db.ForeignKey('orden_compra.id'), nullable=False)
    producto_id = db.Column(db.Integer, db.ForeignKey('producto.id'), nullable=False)
    cantidad = db.Column(db.Integer, nullable=False)
    
    orden = db.relationship('OrdenCompra', backref='items')
    producto = db.relationship('Producto', backref='items_orden')
    
    def __repr__(self):
        return f'<ItemOrden {self.id}>'

# Rutas de la aplicación
@app.route('/')
def index():
    productos = Producto.query.all()
    current_user = get_current_user()
    # Los usuarios tipo 'cliente' solo pueden ver el listado de productos
    return render_template('index.html', productos=productos, current_user=current_user)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        user = Usuario.query.filter_by(email=email).first()
        
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['user_categoria'] = user.categoria
            flash(f'Bienvenido, {user.nombre}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Email o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión exitosamente', 'success')
    return redirect(url_for('index'))

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        
        # Validar contraseña
        is_valid, password_msg = validar_contraseña(password)
        if not is_valid:
            flash(password_msg, 'error')
            return render_template('registro.html')
        
        # Verificar que las contraseñas coincidan
        passwords_match, match_msg = verificar_contraseñas(password, confirm_password)
        if not passwords_match:
            flash(match_msg, 'error')
            return render_template('registro.html')
        
        # Verificar si el email ya existe
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado!', 'error')
            return render_template('registro.html')
        
        # Crear usuario como cliente por defecto
        usuario = Usuario(
            nombre=nombre,
            email=email,
            categoria='cliente'  # Los registros externos son clientes por defecto
        )
        usuario.set_password(password)
        
        try:
            db.session.add(usuario)
            db.session.flush()  # Para obtener el ID
            
            # Finalizar el hash con el ID si es un usuario nuevo
            if usuario.finalize_password():
                db.session.commit()
            else:
                # Si no había contraseña temporal, hacer commit normal
                db.session.commit()
            
            flash('Usuario registrado exitosamente! Ahora puedes iniciar sesión.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar usuario: {str(e)}', 'error')
    
    return render_template('registro.html')

@app.route('/agregar', methods=['GET', 'POST'])
@login_required
@role_required(['supervisor', 'gerente'])
def agregar_producto():
    if request.method == 'POST':
        nombre = request.form['nombre']
        cantidad = int(request.form['cantidad'])
        tipo_material = request.form['tipo_material']
        precio = float(request.form.get('precio', 1000.0))
        descripcion = request.form['descripcion']
        proveedor = request.form['proveedor']
        
        producto = Producto(
            nombre=nombre,
            cantidad=cantidad,
            tipo_material=tipo_material,
            precio=precio,
            descripcion=descripcion,
            proveedor=proveedor
        )
        
        try:
            db.session.add(producto)
            db.session.commit()
            flash('Producto agregado exitosamente!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error al agregar producto: {str(e)}', 'error')
    
    return render_template('agregar.html')

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['supervisor', 'gerente'])
def editar_producto(id):
    producto = Producto.query.get_or_404(id)
    
    if request.method == 'POST':
        producto.nombre = request.form['nombre']
        producto.cantidad = int(request.form['cantidad'])
        producto.tipo_material = request.form['tipo_material']
        producto.precio = float(request.form.get('precio', producto.precio))
        producto.descripcion = request.form['descripcion']
        producto.proveedor = request.form['proveedor']
        
        try:
            db.session.commit()
            flash('Producto actualizado exitosamente!', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            flash(f'Error al actualizar producto: {str(e)}', 'error')
    
    return render_template('editar.html', producto=producto)

@app.route('/eliminar/<int:id>')
@login_required
@role_required(['supervisor', 'gerente'])
def eliminar_producto(id):
    producto = Producto.query.get_or_404(id)
    
    try:
        db.session.delete(producto)
        db.session.commit()
        flash('Producto eliminado exitosamente!', 'success')
    except Exception as e:
        flash(f'Error al eliminar producto: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/ver/<int:id>')
def ver_producto(id):
    producto = Producto.query.get_or_404(id)
    return render_template('ver.html', producto=producto)

# Rutas para Usuarios
@app.route('/usuarios')
@login_required
@role_required(['supervisor', 'gerente'])
def usuarios():
    usuarios = Usuario.query.all()
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuarios/agregar', methods=['GET', 'POST'])
@login_required
@role_required(['gerente'])
def agregar_usuario():
    if request.method == 'POST':
        nombre = request.form['nombre']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        categoria = request.form['categoria']
        
        # Validar contraseña
        is_valid, password_msg = validar_contraseña(password)
        if not is_valid:
            flash(password_msg, 'error')
            return render_template('agregar_usuario.html')
        
        # Verificar que las contraseñas coincidan
        passwords_match, match_msg = verificar_contraseñas(password, confirm_password)
        if not passwords_match:
            flash(match_msg, 'error')
            return render_template('agregar_usuario.html')
        
        # Verificar si el email ya existe
        if Usuario.query.filter_by(email=email).first():
            flash('El email ya está registrado!', 'error')
            return render_template('agregar_usuario.html')
        
        usuario = Usuario(
            nombre=nombre,
            email=email,
            categoria=categoria
        )
        usuario.set_password(password)
        
        try:
            db.session.add(usuario)
            db.session.flush()  # Para obtener el ID sin hacer commit
            
            # Finalizar el hash con el ID si es un usuario nuevo
            if usuario.finalize_password():
                db.session.flush()
            else:
                db.session.flush()
            
            # Si la categoría es 'cliente', crear automáticamente un Cliente asociado
            if categoria == 'cliente':
                identificador = request.form.get('identificador', email)  # Usar email como identificador por defecto si no se proporciona
                tipo_persona = request.form.get('tipo_persona', 'fisica')
                direccion = request.form.get('direccion', '')
                telefono = request.form.get('telefono', '')
                mail = request.form.get('mail_cliente', email)
                datos_adicionales = request.form.get('datos_adicionales', '')
                
                # Verificar si el identificador ya existe
                if Cliente.query.filter_by(identificador=identificador).first():
                    db.session.rollback()
                    flash(f'El identificador {identificador} ya está registrado!', 'error')
                    return render_template('agregar_usuario.html')
                
                cliente = Cliente(
                    identificador=identificador,
                    tipo_persona=tipo_persona,
                    nombre=nombre,
                    direccion=direccion,
                    telefono=telefono,
                    mail=mail,
                    datos_adicionales=datos_adicionales,
                    usuario_id=usuario.id
                )
                db.session.add(cliente)
            
            db.session.commit()
            flash('Usuario agregado exitosamente!', 'success')
            if categoria == 'cliente':
                flash('Cliente asociado creado automáticamente.', 'success')
            return redirect(url_for('usuarios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar usuario: {str(e)}', 'error')
    
    return render_template('agregar_usuario.html')

@app.route('/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['gerente'])
def editar_usuario(id):
    usuario = Usuario.query.get_or_404(id)
    
    if request.method == 'POST':
        usuario.nombre = request.form['nombre']
        usuario.email = request.form['email']
        nueva_categoria = request.form['categoria']
        categoria_anterior = usuario.categoria
        usuario.categoria = nueva_categoria
        
        # Verificar si el email ya existe en otro usuario
        existing_user = Usuario.query.filter_by(email=usuario.email).first()
        if existing_user and existing_user.id != usuario.id:
            flash('El email ya está registrado por otro usuario!', 'error')
            return render_template('editar_usuario.html', usuario=usuario)
        
        # Si se proporciona una nueva contraseña, validarla y actualizarla
        new_password = request.form['password']
        if new_password:
            # Validar contraseña
            is_valid, password_msg = validar_contraseña(new_password)
            if not is_valid:
                flash(password_msg, 'error')
                return render_template('editar_usuario.html', usuario=usuario)
            
            usuario.set_password(new_password)
        
        # Manejar cambios de categoría relacionados con Cliente
        if categoria_anterior == 'cliente' and nueva_categoria != 'cliente':
            # Si el usuario cambia DE 'cliente' a otra categoría, eliminar el Cliente asociado
            if usuario.cliente:
                db.session.delete(usuario.cliente)
                flash('Registro de Cliente eliminado al cambiar la categoría.', 'info')
        
        elif nueva_categoria == 'cliente':
            # Si el usuario cambia A 'cliente'
            if usuario.cliente:
                # Si ya tiene cliente asociado, actualizar sus datos
                cliente = usuario.cliente
                cliente.nombre = usuario.nombre
                cliente.identificador = request.form.get('identificador', cliente.identificador)
                cliente.tipo_persona = request.form.get('tipo_persona', cliente.tipo_persona)
                cliente.direccion = request.form.get('direccion', cliente.direccion)
                cliente.telefono = request.form.get('telefono', cliente.telefono)
                cliente.mail = request.form.get('mail_cliente', usuario.email)
                cliente.datos_adicionales = request.form.get('datos_adicionales', cliente.datos_adicionales)
            else:
                # Si no tiene cliente asociado, crear uno nuevo
                identificador = request.form.get('identificador', usuario.email)
                tipo_persona = request.form.get('tipo_persona', 'fisica')
                
                # Verificar si el identificador ya existe
                cliente_existente = Cliente.query.filter_by(identificador=identificador).first()
                if cliente_existente and cliente_existente.usuario_id != usuario.id:
                    flash(f'El identificador {identificador} ya está registrado!', 'error')
                    return render_template('editar_usuario.html', usuario=usuario)
                
                cliente = Cliente(
                    identificador=identificador,
                    tipo_persona=tipo_persona,
                    nombre=usuario.nombre,
                    direccion=request.form.get('direccion', ''),
                    telefono=request.form.get('telefono', ''),
                    mail=request.form.get('mail_cliente', usuario.email),
                    datos_adicionales=request.form.get('datos_adicionales', ''),
                    usuario_id=usuario.id
                )
                db.session.add(cliente)
                flash('Cliente asociado creado automáticamente.', 'success')
        
        try:
            db.session.commit()
            flash('Usuario actualizado exitosamente!', 'success')
            return redirect(url_for('usuarios'))
        except Exception as e:
            flash(f'Error al actualizar usuario: {str(e)}', 'error')
    
    return render_template('editar_usuario.html', usuario=usuario)

@app.route('/usuarios/eliminar/<int:id>')
@login_required
@role_required(['gerente'])
def eliminar_usuario(id):
    usuario = Usuario.query.get_or_404(id)
    
    try:
        db.session.delete(usuario)
        db.session.commit()
        flash('Usuario eliminado exitosamente!', 'success')
    except Exception as e:
        flash(f'Error al eliminar usuario: {str(e)}', 'error')
    
    return redirect(url_for('usuarios'))

@app.route('/usuarios/ver/<int:id>')
@login_required
@role_required(['supervisor', 'gerente'])
def ver_usuario(id):
    usuario = Usuario.query.get_or_404(id)
    return render_template('ver_usuario.html', usuario=usuario)

# Rutas para Clientes
@app.route('/clientes')
@login_required
@role_required(['supervisor', 'gerente'])
def clientes():
    # Mostrar clientes que:
    # 1. No tienen usuario asociado (agregados directamente)
    # 2. Tienen usuario asociado con categoría 'cliente'
    clientes = Cliente.query.filter(
        or_(
            Cliente.usuario_id.is_(None),
            Cliente.usuario_id.in_(
                db.session.query(Usuario.id).filter(Usuario.categoria == 'cliente')
            )
        )
    ).all()
    return render_template('clientes.html', clientes=clientes)

@app.route('/clientes/agregar', methods=['GET', 'POST'])
@login_required
@role_required(['supervisor', 'gerente'])
def agregar_cliente():
    if request.method == 'POST':
        identificador = request.form['identificador']
        tipo_persona = request.form['tipo_persona']
        nombre = request.form['nombre']
        direccion = request.form.get('direccion', '')
        telefono = request.form.get('telefono', '')
        mail = request.form.get('mail', '')
        datos_adicionales = request.form.get('datos_adicionales', '')
        
        # Verificar si el identificador ya existe
        if Cliente.query.filter_by(identificador=identificador).first():
            flash('El identificador ya está registrado!', 'error')
            return render_template('agregar_cliente.html')
        
        cliente = Cliente(
            identificador=identificador,
            tipo_persona=tipo_persona,
            nombre=nombre,
            direccion=direccion,
            telefono=telefono,
            mail=mail,
            datos_adicionales=datos_adicionales
        )
        
        try:
            db.session.add(cliente)
            db.session.commit()
            flash('Cliente agregado exitosamente!', 'success')
            return redirect(url_for('clientes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al agregar cliente: {str(e)}', 'error')
    
    return render_template('agregar_cliente.html')

@app.route('/clientes/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['supervisor', 'gerente'])
def editar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    
    if request.method == 'POST':
        identificador = request.form['identificador']
        tipo_persona = request.form['tipo_persona']
        nombre = request.form['nombre']
        direccion = request.form.get('direccion', '')
        telefono = request.form.get('telefono', '')
        mail = request.form.get('mail', '')
        datos_adicionales = request.form.get('datos_adicionales', '')
        
        # Verificar si el identificador ya existe en otro cliente
        existing_cliente = Cliente.query.filter_by(identificador=identificador).first()
        if existing_cliente and existing_cliente.id != cliente.id:
            flash('El identificador ya está registrado por otro cliente!', 'error')
            return render_template('editar_cliente.html', cliente=cliente)
        
        cliente.identificador = identificador
        cliente.tipo_persona = tipo_persona
        cliente.nombre = nombre
        cliente.direccion = direccion
        cliente.telefono = telefono
        cliente.mail = mail
        cliente.datos_adicionales = datos_adicionales
        
        try:
            db.session.commit()
            flash('Cliente actualizado exitosamente!', 'success')
            return redirect(url_for('clientes'))
        except Exception as e:
            flash(f'Error al actualizar cliente: {str(e)}', 'error')
    
    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/clientes/eliminar/<int:id>')
@login_required
@role_required(['supervisor', 'gerente'])
def eliminar_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    
    try:
        db.session.delete(cliente)
        db.session.commit()
        flash('Cliente eliminado exitosamente!', 'success')
    except Exception as e:
        flash(f'Error al eliminar cliente: {str(e)}', 'error')
    
    return redirect(url_for('clientes'))

@app.route('/clientes/ver/<int:id>')
@login_required
@role_required(['supervisor', 'gerente'])
def ver_cliente(id):
    cliente = Cliente.query.get_or_404(id)
    return render_template('ver_cliente.html', cliente=cliente)

# Funciones auxiliares para el carrito
def get_carrito():
    """Obtiene el carrito de la sesión"""
    if 'carrito' not in session:
        session['carrito'] = []
    return session['carrito']

def calcular_total_carrito():
    """Calcula el total del carrito"""
    carrito = get_carrito()
    total = 0
    for item in carrito:
        producto = Producto.query.get(item['producto_id'])
        if producto:
            precio_unitario = producto.precio
            total += precio_unitario * item['cantidad']
    return total

# Rutas para Carrito de Compras (solo clientes)
@app.route('/carrito/agregar/<int:producto_id>', methods=['POST'])
@login_required
@role_required(['cliente'])
def agregar_al_carrito(producto_id):
    producto = Producto.query.get_or_404(producto_id)
    cantidad = int(request.form.get('cantidad', 1))
    
    if cantidad <= 0:
        flash('La cantidad debe ser mayor a cero', 'error')
        return redirect(url_for('index'))
    
    if cantidad > producto.cantidad:
        flash(f'No hay suficiente stock. Stock disponible: {producto.cantidad}', 'error')
        return redirect(url_for('index'))
    
    carrito = get_carrito()
    
    # Verificar si el producto ya está en el carrito
    for item in carrito:
        if item['producto_id'] == producto_id:
            nueva_cantidad = item['cantidad'] + cantidad
            if nueva_cantidad > producto.cantidad:
                flash(f'No hay suficiente stock. Stock disponible: {producto.cantidad}', 'error')
                return redirect(url_for('ver_carrito'))
            item['cantidad'] = nueva_cantidad
            session.modified = True
            flash(f'Cantidad actualizada en el carrito', 'success')
            return redirect(url_for('ver_carrito'))
    
    # Agregar nuevo producto al carrito
    carrito.append({
        'producto_id': producto_id,
        'cantidad': cantidad
    })
    session.modified = True
    flash(f'{producto.nombre} agregado al carrito', 'success')
    return redirect(url_for('ver_carrito'))

@app.route('/carrito')
@login_required
@role_required(['cliente'])
def ver_carrito():
    carrito = get_carrito()
    items_carrito = []
    for item in carrito:
        producto = Producto.query.get(item['producto_id'])
        if producto:
            precio_unitario = producto.precio
            items_carrito.append({
                'producto': producto,
                'cantidad': item['cantidad'],
                'subtotal': precio_unitario * item['cantidad']
            })
    
    total = calcular_total_carrito()
    return render_template('carrito.html', items_carrito=items_carrito, total=total)

@app.route('/carrito/eliminar/<int:producto_id>')
@login_required
@role_required(['cliente'])
def eliminar_del_carrito(producto_id):
    carrito = get_carrito()
    carrito[:] = [item for item in carrito if item['producto_id'] != producto_id]
    session.modified = True
    flash('Producto eliminado del carrito', 'success')
    return redirect(url_for('ver_carrito'))

@app.route('/carrito/actualizar/<int:producto_id>', methods=['POST'])
@login_required
@role_required(['cliente'])
def actualizar_cantidad_carrito(producto_id):
    cantidad = int(request.form.get('cantidad', 1))
    producto = Producto.query.get_or_404(producto_id)
    
    if cantidad <= 0:
        flash('La cantidad debe ser mayor a cero', 'error')
        return redirect(url_for('ver_carrito'))
    
    if cantidad > producto.cantidad:
        flash(f'No hay suficiente stock. Stock disponible: {producto.cantidad}', 'error')
        return redirect(url_for('ver_carrito'))
    
    carrito = get_carrito()
    for item in carrito:
        if item['producto_id'] == producto_id:
            item['cantidad'] = cantidad
            session.modified = True
            flash('Cantidad actualizada', 'success')
            break
    
    return redirect(url_for('ver_carrito'))

# Ruta para Checkout
@app.route('/checkout', methods=['GET', 'POST'])
@login_required
@role_required(['cliente'])
def checkout():
    carrito = get_carrito()
    
    if not carrito:
        flash('El carrito está vacío', 'error')
        return redirect(url_for('index'))
    
    current_user = get_current_user()
    cliente = Cliente.query.filter_by(usuario_id=current_user.id).first()
    
    if not cliente:
        flash('No se encontró información del cliente', 'error')
        return redirect(url_for('index'))
    
    items_carrito = []
    for item in carrito:
        producto = Producto.query.get(item['producto_id'])
        if producto:
            precio_unitario = producto.precio
            items_carrito.append({
                'producto': producto,
                'cantidad': item['cantidad'],
                'subtotal': precio_unitario * item['cantidad']
            })
    
    total = calcular_total_carrito()
    
    if request.method == 'POST':
        direccion_entrega = request.form.get('direccion_entrega', '').strip()
        
        if not direccion_entrega:
            flash('Debe especificar una dirección de entrega', 'error')
            return render_template('checkout.html', items_carrito=items_carrito, total=total, cliente=cliente)
        
        # Generar identificador único para la orden
        ultima_orden = OrdenCompra.query.order_by(OrdenCompra.identificador.desc()).first()
        nuevo_identificador = (ultima_orden.identificador + 1) if ultima_orden else 1000
        
        # Crear la orden
        orden = OrdenCompra(
            identificador=nuevo_identificador,
            cliente_id=cliente.id,
            direccion_entrega=direccion_entrega,
            estado='pendiente_confirmacion'
        )
        db.session.add(orden)
        db.session.flush()
        
        # Crear los items de la orden
        for item in carrito:
            producto = Producto.query.get(item['producto_id'])
            if producto:
                item_orden = ItemOrden(
                    orden_id=orden.id,
                    producto_id=producto.id,
                    cantidad=item['cantidad']
                )
                db.session.add(item_orden)
                
                # Reducir stock del producto
                producto.cantidad -= item['cantidad']
        
        # Limpiar el carrito
        session['carrito'] = []
        session.modified = True
        
        try:
            db.session.commit()
            flash(f'Orden #{orden.identificador} creada exitosamente!', 'success')
            return redirect(url_for('mis_ordenes'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la orden: {str(e)}', 'error')
    
    return render_template('checkout.html', items_carrito=items_carrito, total=total, cliente=cliente)

# Rutas para Clientes - Mis Órdenes
@app.route('/mis-ordenes')
@login_required
@role_required(['cliente'])
def mis_ordenes():
    current_user = get_current_user()
    cliente = Cliente.query.filter_by(usuario_id=current_user.id).first()
    
    if not cliente:
        flash('No se encontró información del cliente', 'error')
        return redirect(url_for('index'))
    
    ordenes = OrdenCompra.query.filter_by(cliente_id=cliente.id).order_by(OrdenCompra.fecha.desc()).all()
    return render_template('mis_ordenes.html', ordenes=ordenes)

@app.route('/ordenes/ver/<int:id>')
@login_required
def ver_orden(id):
    orden = OrdenCompra.query.get_or_404(id)
    current_user = get_current_user()
    
    # Verificar permisos
    if current_user.categoria == 'cliente':
        cliente = Cliente.query.filter_by(usuario_id=current_user.id).first()
        if not cliente or orden.cliente_id != cliente.id:
            flash('No tienes permisos para ver esta orden', 'error')
            return redirect(url_for('index'))
    elif current_user.categoria not in ['supervisor', 'gerente']:
        flash('No tienes permisos para ver esta orden', 'error')
        return redirect(url_for('index'))
    
    items = ItemOrden.query.filter_by(orden_id=orden.id).all()
    total = orden.calcular_total()
    
    return render_template('ver_orden.html', orden=orden, items=items, total=total, current_user=current_user)

@app.route('/ordenes/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required(['cliente'])
def editar_orden(id):
    orden = OrdenCompra.query.get_or_404(id)
    current_user = get_current_user()
    cliente = Cliente.query.filter_by(usuario_id=current_user.id).first()
    
    if not cliente or orden.cliente_id != cliente.id:
        flash('No tienes permisos para editar esta orden', 'error')
        return redirect(url_for('mis_ordenes'))
    
    if orden.estado != 'pendiente_confirmacion':
        flash('Solo se pueden editar órdenes con estado "Pendiente de Confirmación"', 'error')
        return redirect(url_for('ver_orden', id=orden.id))
    
    items = ItemOrden.query.filter_by(orden_id=orden.id).all()
    
    if request.method == 'POST':
        # Actualizar dirección de entrega
        orden.direccion_entrega = request.form.get('direccion_entrega', '').strip()
        
        # Actualizar cantidades de productos
        for item in items:
            nueva_cantidad = int(request.form.get(f'cantidad_{item.id}', item.cantidad))
            if nueva_cantidad > 0:
                # Verificar stock disponible
                producto = item.producto
                cantidad_original = item.cantidad
                diferencia = nueva_cantidad - cantidad_original
                
                # El stock disponible incluye lo que está reservado en esta orden
                stock_disponible_real = producto.cantidad + cantidad_original
                
                if nueva_cantidad > stock_disponible_real:
                    flash(f'No hay suficiente stock de {producto.nombre}. Stock disponible: {stock_disponible_real}', 'error')
                    return redirect(url_for('editar_orden', id=orden.id))
                
                # Restaurar cantidad original al stock y luego restar la nueva cantidad
                producto.cantidad += cantidad_original
                producto.cantidad -= nueva_cantidad
                item.cantidad = nueva_cantidad
        
        try:
            db.session.commit()
            flash('Orden actualizada exitosamente', 'success')
            return redirect(url_for('ver_orden', id=orden.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar la orden: {str(e)}', 'error')
    
    total = orden.calcular_total()
    return render_template('editar_orden.html', orden=orden, items=items, total=total)

@app.route('/ordenes/anular/<int:id>')
@login_required
@role_required(['cliente'])
def anular_orden(id):
    orden = OrdenCompra.query.get_or_404(id)
    current_user = get_current_user()
    cliente = Cliente.query.filter_by(usuario_id=current_user.id).first()
    
    if not cliente or orden.cliente_id != cliente.id:
        flash('No tienes permisos para anular esta orden', 'error')
        return redirect(url_for('mis_ordenes'))
    
    if orden.estado in ['completada', 'cancelada']:
        mensaje = 'No se pueden anular órdenes completadas' if orden.estado == 'completada' else 'Esta orden ya fue cancelada'
        flash(mensaje, 'error')
        return redirect(url_for('ver_orden', id=orden.id))
    
    # Restaurar stock
    items = ItemOrden.query.filter_by(orden_id=orden.id).all()
    for item in items:
        item.producto.cantidad += item.cantidad
    
    orden.estado = 'cancelada'
    
    try:
        db.session.commit()
        flash('Orden anulada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al anular la orden: {str(e)}', 'error')
    
    return redirect(url_for('mis_ordenes'))

# Rutas para Supervisor/Gerente - Todas las Órdenes
@app.route('/ordenes')
@login_required
@role_required(['supervisor', 'gerente'])
def todas_las_ordenes():
    ordenes = OrdenCompra.query.order_by(OrdenCompra.fecha.desc()).all()
    return render_template('todas_las_ordenes.html', ordenes=ordenes)

@app.route('/ordenes/cambiar-estado/<int:id>', methods=['POST'])
@login_required
@role_required(['supervisor', 'gerente'])
def cambiar_estado_orden(id):
    orden = OrdenCompra.query.get_or_404(id)
    nuevo_estado = request.form.get('estado')
    
    estados_validos = ['pendiente_confirmacion', 'confirmada', 'en_transito', 'completada', 'cancelada']
    
    if nuevo_estado not in estados_validos:
        flash('Estado inválido', 'error')
        return redirect(url_for('ver_orden', id=orden.id))
    
    estados_finales = {'completada', 'cancelada'}
    if orden.estado in estados_finales:
        flash('Las órdenes completadas o canceladas no pueden modificarse', 'error')
        return redirect(url_for('ver_orden', id=orden.id))
    
    orden_anterior = orden.estado
    orden.estado = nuevo_estado
    
    # Si se cancela, restaurar stock
    if nuevo_estado == 'cancelada' and orden_anterior != 'cancelada':
        items = ItemOrden.query.filter_by(orden_id=orden.id).all()
        for item in items:
            item.producto.cantidad += item.cantidad
    
    try:
        db.session.commit()
        flash(f'Estado de la orden cambiado a "{nuevo_estado.replace("_", " ").title()}"', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar el estado: {str(e)}', 'error')
    
    return redirect(url_for('ver_orden', id=orden.id))

# Ruta para generar reportes de órdenes
@app.route('/ordenes/generar-reporte', methods=['GET'])
@login_required
@role_required(['supervisor', 'gerente'])
def generar_reporte_ordenes():
    # Obtener parámetros de filtro
    formato = request.args.get('formato', 'excel')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    monto_minimo = request.args.get('monto_minimo', '')
    monto_maximo = request.args.get('monto_maximo', '')
    
    # Construir query base
    query = OrdenCompra.query
    
    # Aplicar filtros de fecha
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(OrdenCompra.fecha >= fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            # Incluir todo el día
            fecha_hasta_obj = fecha_hasta_obj.replace(hour=23, minute=59, second=59)
            query = query.filter(OrdenCompra.fecha <= fecha_hasta_obj)
        except ValueError:
            pass
    
    # Obtener todas las órdenes para calcular totales y aplicar filtro de monto
    ordenes = query.order_by(OrdenCompra.fecha.desc()).all()
    
    # Aplicar filtros de monto
    ordenes_filtradas = []
    for orden in ordenes:
        total = orden.calcular_total()
        
        if monto_minimo:
            try:
                if total < float(monto_minimo):
                    continue
            except ValueError:
                pass
        
        if monto_maximo:
            try:
                if total > float(monto_maximo):
                    continue
            except ValueError:
                pass
        
        ordenes_filtradas.append(orden)
    
    # Generar reporte según el formato
    if formato == 'excel':
        return generar_reporte_excel(ordenes_filtradas, fecha_desde, fecha_hasta, monto_minimo, monto_maximo)
    elif formato == 'pdf':
        return generar_reporte_pdf(ordenes_filtradas, fecha_desde, fecha_hasta, monto_minimo, monto_maximo)
    else:
        flash('Formato de reporte no válido', 'error')
        return redirect(url_for('todas_las_ordenes'))

def generar_reporte_excel(ordenes, fecha_desde, fecha_hasta, monto_minimo, monto_maximo):
    """Genera un reporte en formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Órdenes"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "REPORTE DE ÓRDENES DE COMPRA"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información de filtros
    row = 3
    ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    row += 1
    
    filtros = []
    if fecha_desde:
        filtros.append(f"Fecha desde: {fecha_desde}")
    if fecha_hasta:
        filtros.append(f"Fecha hasta: {fecha_hasta}")
    if monto_minimo:
        filtros.append(f"Monto mínimo: ${monto_minimo}")
    if monto_maximo:
        filtros.append(f"Monto máximo: ${monto_maximo}")
    
    if filtros:
        ws[f'A{row}'] = "Filtros aplicados: " + " | ".join(filtros)
    else:
        ws[f'A{row}'] = "Sin filtros aplicados (todas las órdenes)"
    row += 2
    
    # Encabezados
    headers = ['# Orden', 'Cliente', 'Fecha', 'Estado', 'Total', 'Dirección Entrega', 'Items']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row += 1
    total_general = 0
    for orden in ordenes:
        total_orden = orden.calcular_total()
        total_general += total_orden
        
        # Obtener items de la orden
        items = ItemOrden.query.filter_by(orden_id=orden.id).all()
        items_texto = ", ".join([f"{item.producto.nombre} (x{item.cantidad})" for item in items])
        
        # Estado formateado
        estado_map = {
            'pendiente_confirmacion': 'Pendiente de Confirmación',
            'confirmada': 'Confirmada',
            'en_transito': 'En Tránsito',
            'completada': 'Completada',
            'cancelada': 'Cancelada'
        }
        estado_texto = estado_map.get(orden.estado, orden.estado)
        
        data = [
            orden.identificador,
            orden.cliente.nombre,
            orden.fecha.strftime('%d/%m/%Y %H:%M'),
            estado_texto,
            f"${total_orden:.2f}",
            orden.direccion_entrega or 'N/A',
            items_texto[:100] + '...' if len(items_texto) > 100 else items_texto
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 5:  # Columna de Total
                cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Total general
    row += 1
    ws.cell(row=row, column=4, value="TOTAL GENERAL:").font = Font(bold=True)
    ws.cell(row=row, column=5, value=f"${total_general:.2f}").font = Font(bold=True)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 40
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_ordenes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def generar_reporte_pdf(ordenes, fecha_desde, fecha_hasta, monto_minimo, monto_maximo):
    """Genera un reporte en formato PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenedor para elementos del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Center
        spaceAfter=30
    )
    title = Paragraph("REPORTE DE ÓRDENES DE COMPRA", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información de filtros
    info_text = f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>"
    filtros = []
    if fecha_desde:
        filtros.append(f"Fecha desde: {fecha_desde}")
    if fecha_hasta:
        filtros.append(f"Fecha hasta: {fecha_hasta}")
    if monto_minimo:
        filtros.append(f"Monto mínimo: ${monto_minimo}")
    if monto_maximo:
        filtros.append(f"Monto máximo: ${monto_maximo}")
    
    if filtros:
        info_text += f"<b>Filtros aplicados:</b> {' | '.join(filtros)}"
    else:
        info_text += "<b>Sin filtros aplicados</b> (todas las órdenes)"
    
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['# Orden', 'Cliente', 'Fecha', 'Estado', 'Total']]
    
    total_general = 0
    for orden in ordenes:
        total_orden = orden.calcular_total()
        total_general += total_orden
        
        estado_map = {
            'pendiente_confirmacion': 'Pendiente',
            'confirmada': 'Confirmada',
            'en_transito': 'En Tránsito',
            'completada': 'Completada',
            'cancelada': 'Cancelada'
        }
        estado_texto = estado_map.get(orden.estado, orden.estado)
        
        data.append([
            str(orden.identificador),
            orden.cliente.nombre[:30] + '...' if len(orden.cliente.nombre) > 30 else orden.cliente.nombre,
            orden.fecha.strftime('%d/%m/%Y'),
            estado_texto,
            f"${total_orden:.2f}"
        ])
    
    # Agregar total general (fila final compacta para evitar superposición visual)
    data.append(['TOTALES', '', '', '', f'${total_general:.2f}'])
    
    # Crear tabla
    table = Table(data, colWidths=[1*inch, 2.5*inch, 1.2*inch, 1.3*inch, 1*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 1), (4, -2), 'RIGHT'),  # Alinear totales a la derecha
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        
        # Fila de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (0, -1), (0, -1), 'CENTER'),
        ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_ordenes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

# Ruta para generar reportes de productos
@app.route('/productos/generar-reporte', methods=['GET'])
@login_required
@role_required(['supervisor', 'gerente'])
def generar_reporte_productos():
    # Obtener parámetros de filtro
    formato = request.args.get('formato', 'excel')
    stock_minimo = request.args.get('stock_minimo', '')
    stock_maximo = request.args.get('stock_maximo', '')
    precio_minimo = request.args.get('precio_minimo', '')
    precio_maximo = request.args.get('precio_maximo', '')
    
    # Construir query base
    query = Producto.query
    
    # Aplicar filtros de stock
    if stock_minimo:
        try:
            query = query.filter(Producto.cantidad >= int(stock_minimo))
        except ValueError:
            pass
    
    if stock_maximo:
        try:
            query = query.filter(Producto.cantidad <= int(stock_maximo))
        except ValueError:
            pass
    
    # Obtener todos los productos para calcular precios y aplicar filtro de precio
    productos = query.order_by(Producto.nombre).all()
    
    # Aplicar filtros de precio (precio fijo por unidad)
    productos_filtrados = []
    for producto in productos:
        precio = producto.precio
        
        if precio_minimo:
            try:
                if precio < float(precio_minimo):
                    continue
            except ValueError:
                pass
        
        if precio_maximo:
            try:
                if precio > float(precio_maximo):
                    continue
            except ValueError:
                pass
        
        productos_filtrados.append(producto)
    
    # Generar reporte según el formato
    if formato == 'excel':
        return generar_reporte_productos_excel(productos_filtrados, stock_minimo, stock_maximo, precio_minimo, precio_maximo)
    elif formato == 'pdf':
        return generar_reporte_productos_pdf(productos_filtrados, stock_minimo, stock_maximo, precio_minimo, precio_maximo)
    else:
        flash('Formato de reporte no válido', 'error')
        return redirect(url_for('index'))

def generar_reporte_productos_excel(productos, stock_minimo, stock_maximo, precio_minimo, precio_maximo):
    """Genera un reporte de productos en formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Productos"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "REPORTE DE PRODUCTOS"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información de filtros
    row = 3
    ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    row += 1
    
    filtros = []
    if stock_minimo:
        filtros.append(f"Stock mínimo: {stock_minimo}")
    if stock_maximo:
        filtros.append(f"Stock máximo: {stock_maximo}")
    if precio_minimo:
        filtros.append(f"Precio mínimo: ${precio_minimo}/unidad")
    if precio_maximo:
        filtros.append(f"Precio máximo: ${precio_maximo}/unidad")
    
    if filtros:
        ws[f'A{row}'] = "Filtros aplicados: " + " | ".join(filtros)
    else:
        ws[f'A{row}'] = "Sin filtros aplicados (todos los productos)"
    row += 2
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Descripción', 'Cantidad (Stock)', 'Tipo de Material', 'Precio ($/unidad)', 'Proveedor', 'Fecha Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row += 1
    total_stock = 0
    valor_total_inventario = 0
    
    for producto in productos:
        precio = producto.precio
        valor_producto = precio * producto.cantidad
        total_stock += producto.cantidad
        valor_total_inventario += valor_producto
        
        data = [
            producto.id,
            producto.nombre,
            producto.descripcion or 'N/A',
            producto.cantidad,
            producto.tipo_material,
            f"${precio:.2f}",
            producto.proveedor,
            producto.fecha_creacion.strftime('%d/%m/%Y')
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col == 4:  # Columna de Cantidad
                cell.alignment = Alignment(horizontal='center')
            elif col == 6:  # Columna de Precio
                cell.alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Totales
    row += 1
    ws.cell(row=row, column=3, value="TOTALES:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_stock).font = Font(bold=True)
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    
    row += 1
    ws.cell(row=row, column=3, value="VALOR TOTAL INVENTARIO:").font = Font(bold=True)
    ws.cell(row=row, column=6, value=f"${valor_total_inventario:.2f}").font = Font(bold=True)
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def generar_reporte_productos_pdf(productos, stock_minimo, stock_maximo, precio_minimo, precio_maximo):
    """Genera un reporte de productos en formato PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenedor para elementos del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Center
        spaceAfter=30
    )
    title = Paragraph("REPORTE DE PRODUCTOS", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información de filtros
    info_text = f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>"
    filtros = []
    if stock_minimo:
        filtros.append(f"Stock mínimo: {stock_minimo}")
    if stock_maximo:
        filtros.append(f"Stock máximo: {stock_maximo}")
    if precio_minimo:
        filtros.append(f"Precio mínimo: ${precio_minimo}/unidad")
    if precio_maximo:
        filtros.append(f"Precio máximo: ${precio_maximo}/unidad")
    
    if filtros:
        info_text += f"<b>Filtros aplicados:</b> {' | '.join(filtros)}"
    else:
        info_text += "<b>Sin filtros aplicados</b> (todos los productos)"
    
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['ID', 'Nombre', 'Stock', 'Tipo de Material', 'Precio ($/unidad)', 'Proveedor']]
    
    total_stock = 0
    valor_total_inventario = 0
    
    for producto in productos:
        precio = producto.precio
        valor_producto = precio * producto.cantidad
        total_stock += producto.cantidad
        valor_total_inventario += valor_producto
        
        nombre_corto = producto.nombre[:25] + '...' if len(producto.nombre) > 25 else producto.nombre
        proveedor_corto = producto.proveedor[:20] + '...' if len(producto.proveedor) > 20 else producto.proveedor
        
        data.append([
            str(producto.id),
            nombre_corto,
            str(producto.cantidad),
            producto.tipo_material,
            f"${precio:.2f}",
            proveedor_corto
        ])
    
    # Agregar totales
    data.append(['', '<b>TOTALES</b>', f'<b>{total_stock}</b>', '', f'<b>${valor_total_inventario:.2f}</b>', ''])
    
    # Crear tabla
    table = Table(data, colWidths=[0.6*inch, 2*inch, 0.8*inch, 1*inch, 1.2*inch, 1.4*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.black),
        ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
        ('ALIGN', (2, 1), (2, -2), 'CENTER'),  # Alinear stock al centro
        ('ALIGN', (4, 1), (4, -2), 'RIGHT'),  # Alinear precio a la derecha
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
        
        # Fila de totales
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('ALIGN', (2, -1), (2, -1), 'CENTER'),
        ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

# Ruta para generar reportes de clientes
@app.route('/clientes/generar-reporte', methods=['GET'])
@login_required
@role_required(['supervisor', 'gerente'])
def generar_reporte_clientes():
    # Obtener parámetros
    formato = request.args.get('formato', 'excel')
    
    # Obtener clientes que no tienen usuario asociado o tienen usuario con categoría 'cliente'
    from sqlalchemy import or_
    clientes = Cliente.query.filter(
        or_(
            Cliente.usuario_id.is_(None),
            Cliente.usuario_id.in_(
                db.session.query(Usuario.id).filter(Usuario.categoria == 'cliente')
            )
        )
    ).order_by(Cliente.nombre).all()
    
    # Generar reporte según el formato
    if formato == 'excel':
        return generar_reporte_clientes_excel(clientes)
    elif formato == 'pdf':
        return generar_reporte_clientes_pdf(clientes)
    else:
        flash('Formato de reporte no válido', 'error')
        return redirect(url_for('clientes'))

def generar_reporte_clientes_excel(clientes):
    """Genera un reporte de clientes en formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "REPORTE DE CLIENTES"
    ws['A1'].font = title_font
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información
    row = 3
    ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    row += 1
    ws[f'A{row}'] = f"Total de clientes: {len(clientes)}"
    row += 2
    
    # Encabezados
    headers = ['ID', 'Identificador', 'Nombre', 'Tipo Persona', 'Dirección', 'Teléfono', 'Email', 'Fecha Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row += 1
    for cliente in clientes:
        tipo_persona = 'Persona Física' if cliente.tipo_persona == 'fisica' else 'Persona Jurídica'
        
        data = [
            cliente.id,
            cliente.identificador,
            cliente.nombre,
            tipo_persona,
            cliente.direccion or 'N/A',
            cliente.telefono or 'N/A',
            cliente.mail or 'N/A',
            cliente.fecha_creacion.strftime('%d/%m/%Y')
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
        
        row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def generar_reporte_clientes_pdf(clientes):
    """Genera un reporte de clientes en formato PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenedor para elementos del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Center
        spaceAfter=30
    )
    title = Paragraph("REPORTE DE CLIENTES", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información
    info_text = f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>"
    info_text += f"<b>Total de clientes:</b> {len(clientes)}"
    
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['ID', 'Identificador', 'Nombre', 'Tipo', 'Teléfono', 'Email']]
    
    for cliente in clientes:
        tipo_persona = 'Física' if cliente.tipo_persona == 'fisica' else 'Jurídica'
        nombre_corto = cliente.nombre[:25] + '...' if len(cliente.nombre) > 25 else cliente.nombre
        identificador_corto = cliente.identificador[:20] + '...' if len(cliente.identificador) > 20 else cliente.identificador
        
        data.append([
            str(cliente.id),
            identificador_corto,
            nombre_corto,
            tipo_persona,
            cliente.telefono or '-',
            cliente.mail or '-'
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[0.6*inch, 1.5*inch, 2*inch, 0.8*inch, 1.2*inch, 1.9*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

# Ruta para generar reportes de usuarios
@app.route('/usuarios/generar-reporte', methods=['GET'])
@login_required
@role_required(['supervisor', 'gerente'])
def generar_reporte_usuarios():
    # Obtener parámetros
    formato = request.args.get('formato', 'excel')
    categoria = request.args.get('categoria', '')
    
    # Construir query base
    query = Usuario.query
    
    # Aplicar filtro de categoría si se especifica
    if categoria:
        query = query.filter(Usuario.categoria == categoria)
    
    # Obtener usuarios
    usuarios = query.order_by(Usuario.nombre).all()
    
    # Generar reporte según el formato
    if formato == 'excel':
        return generar_reporte_usuarios_excel(usuarios, categoria)
    elif formato == 'pdf':
        return generar_reporte_usuarios_pdf(usuarios, categoria)
    else:
        flash('Formato de reporte no válido', 'error')
        return redirect(url_for('usuarios'))

def generar_reporte_usuarios_excel(usuarios, categoria):
    """Genera un reporte de usuarios en formato Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Usuarios"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "REPORTE DE USUARIOS"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información
    row = 3
    ws[f'A{row}'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    row += 1
    
    if categoria:
        categoria_map = {
            'gerente': 'Gerente',
            'supervisor': 'Supervisor',
            'usuario_comun': 'Usuario Común',
            'cliente': 'Cliente'
        }
        categoria_texto = categoria_map.get(categoria, categoria)
        ws[f'A{row}'] = f"Filtro aplicado: Categoría = {categoria_texto}"
    else:
        ws[f'A{row}'] = "Sin filtros aplicados (todas las categorías)"
    row += 1
    ws[f'A{row}'] = f"Total de usuarios: {len(usuarios)}"
    row += 2
    
    # Encabezados
    headers = ['ID', 'Nombre', 'Email', 'Categoría', 'Fecha Creación']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    row += 1
    categorias_count = {}
    
    for usuario in usuarios:
        categoria_map = {
            'gerente': 'Gerente',
            'supervisor': 'Supervisor',
            'usuario_comun': 'Usuario Común',
            'cliente': 'Cliente'
        }
        categoria_texto = categoria_map.get(usuario.categoria, usuario.categoria)
        
        # Contar categorías
        if categoria_texto not in categorias_count:
            categorias_count[categoria_texto] = 0
        categorias_count[categoria_texto] += 1
        
        data = [
            usuario.id,
            usuario.nombre,
            usuario.email,
            categoria_texto,
            usuario.fecha_creacion.strftime('%d/%m/%Y')
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
        
        row += 1
    
    # Resumen por categoría
    if not categoria:
        row += 1
        ws.cell(row=row, column=1, value="RESUMEN POR CATEGORÍA:").font = Font(bold=True)
        row += 1
        
        for cat, count in sorted(categorias_count.items()):
            ws.cell(row=row, column=1, value=cat).font = Font(bold=True)
            ws.cell(row=row, column=2, value=count)
            row += 1
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def generar_reporte_usuarios_pdf(usuarios, categoria):
    """Genera un reporte de usuarios en formato PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Contenedor para elementos del PDF
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        alignment=1,  # Center
        spaceAfter=30
    )
    title = Paragraph("REPORTE DE USUARIOS", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Información
    info_text = f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>"
    
    if categoria:
        categoria_map = {
            'gerente': 'Gerente',
            'supervisor': 'Supervisor',
            'usuario_comun': 'Usuario Común',
            'cliente': 'Cliente'
        }
        categoria_texto = categoria_map.get(categoria, categoria)
        info_text += f"<b>Filtro aplicado:</b> Categoría = {categoria_texto}<br/>"
    else:
        info_text += "<b>Sin filtros aplicados</b> (todas las categorías)<br/>"
    
    info_text += f"<b>Total de usuarios:</b> {len(usuarios)}"
    
    info = Paragraph(info_text, styles['Normal'])
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Preparar datos de la tabla
    data = [['ID', 'Nombre', 'Email', 'Categoría']]
    
    categorias_count = {}
    
    for usuario in usuarios:
        categoria_map = {
            'gerente': 'Gerente',
            'supervisor': 'Supervisor',
            'usuario_comun': 'Usuario Común',
            'cliente': 'Cliente'
        }
        categoria_texto = categoria_map.get(usuario.categoria, usuario.categoria)
        
        # Contar categorías
        if categoria_texto not in categorias_count:
            categorias_count[categoria_texto] = 0
        categorias_count[categoria_texto] += 1
        
        nombre_corto = usuario.nombre[:25] + '...' if len(usuario.nombre) > 25 else usuario.nombre
        email_corto = usuario.email[:25] + '...' if len(usuario.email) > 25 else usuario.email
        
        data.append([
            str(usuario.id),
            nombre_corto,
            email_corto,
            categoria_texto
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[0.6*inch, 2.2*inch, 2.2*inch, 1*inch])
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ])
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Agregar resumen por categoría si no hay filtro
    if not categoria and categorias_count:
        elements.append(Spacer(1, 0.3*inch))
        summary_text = "<b>RESUMEN POR CATEGORÍA:</b><br/>"
        for cat, count in sorted(categorias_count.items()):
            summary_text += f"{cat}: {count}<br/>"
        
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear respuesta
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=reporte_usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

def crear_usuario_gerente():
    """Crea el usuario gerente por defecto si no existe"""
    with app.app_context():
        # Verificar si ya existe un gerente
        gerente_existente = Usuario.query.filter_by(email='valeblitochile@gmail.com').first()
        
        if gerente_existente:
            # Actualizar contraseña del gerente existente
            gerente_existente.set_password('SIMA2025')
            try:
                db.session.commit()
                print("[OK] Contraseña del gerente actualizada exitosamente:")
                print(f"   Email: valeblitochile@gmail.com")
                print(f"   Nueva Contraseña: SIMA2025")
                print(f"   Rol: Gerente")
            except Exception as e:
                db.session.rollback()
                print(f"[ERROR] Error al actualizar contraseña del gerente: {str(e)}")
        else:
            gerente = Usuario(
                nombre='Valentina Blito',
                email='valeblitochile@gmail.com',
                categoria='gerente'
            )
            gerente.set_password('SIMA2025')
            
            try:
                db.session.add(gerente)
                db.session.flush()  # Para obtener el ID
                
                # Finalizar el hash con el ID
                if gerente.finalize_password():
                    db.session.commit()
                    print("[OK] Usuario gerente creado exitosamente:")
                    print(f"   Email: valeblitochile@gmail.com")
                    print(f"   Contraseña: SIMA2025")
                    print(f"   Rol: Gerente")
                else:
                    db.session.commit()
                    print("✅ Usuario gerente creado exitosamente")
            except Exception as e:
                db.session.rollback()
                print(f"[ERROR] Error al crear usuario gerente: {str(e)}")

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        crear_usuario_gerente()
    app.run(debug=True)
